from copy import copy
from pathlib import Path

from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.schemas.project import ColumnSchema, ProjectRecord, SheetSchema, TemplateSchema
from app.schemas.run import PaperExtractionResult, RunRecord
from app.storage.repository import Repository, get_repository


class TemplateService:
    def __init__(self, repository: Repository) -> None:
        self.repository = repository

    def parse_template_schema(self, template_path: Path, workbook_filename: str) -> TemplateSchema:
        workbook = load_workbook(
            filename=template_path,
            keep_vba=template_path.suffix.lower() == ".xlsm",
        )
        sheets: list[SheetSchema] = []

        for worksheet in workbook.worksheets:
            header_row_index = self._find_header_row_index(worksheet)
            if header_row_index is None:
                continue

            columns: list[ColumnSchema] = []
            for column_index, cell in enumerate(worksheet[header_row_index], start=1):
                value = cell.value
                if value is None or str(value).strip() == "":
                    continue
                columns.append(
                    ColumnSchema(
                        name=str(value).strip(),
                        column_index=column_index,
                        excel_column=get_column_letter(column_index),
                    )
                )

            if columns:
                sheets.append(
                    SheetSchema(
                        name=worksheet.title,
                        header_row_index=header_row_index,
                        data_start_row_index=header_row_index + 1,
                        columns=columns,
                    )
                )

        if not sheets:
            raise ValueError("No sheet with a non-empty header row was detected in the template workbook.")

        return TemplateSchema(workbook_filename=workbook_filename, sheets=sheets)

    def write_run_workbook(self, project: ProjectRecord, run: RunRecord) -> Path:
        if run.template_stored_filename_snapshot:
            run_template_path = self.repository.run_template_path(
                project.id,
                run.id,
                run.template_stored_filename_snapshot,
            )
            template_path = run_template_path if run_template_path.exists() else self.repository.template_path(project.id, project.template_stored_filename)
        else:
            template_path = self.repository.template_path(project.id, project.template_stored_filename)
        workbook = load_workbook(
            filename=template_path,
            keep_vba=template_path.suffix.lower() == ".xlsm",
        )
        completed_results = [result for result in run.paper_results if result.status.value == "completed"]

        for sheet_schema in run.template_schema_snapshot.sheets:
            worksheet = workbook[sheet_schema.name]
            start_row = self._find_first_empty_row(worksheet, sheet_schema)

            for offset, result in enumerate(completed_results):
                target_row = start_row + offset
                if target_row != sheet_schema.data_start_row_index:
                    self._copy_row_style(worksheet, sheet_schema.data_start_row_index, target_row, sheet_schema.columns)

                sheet_output = result.normalized_output.get(sheet_schema.name, {})
                for column in sheet_schema.columns:
                    worksheet.cell(
                        row=target_row,
                        column=column.column_index,
                        value=self._resolve_cell_value(column.name, result, sheet_output),
                    )

        workbook_filename = f"{run.id}{template_path.suffix.lower() or '.xlsx'}"
        workbook_path = self.repository.workbook_path(project.id, workbook_filename)
        workbook.save(workbook_path)
        return workbook_path

    def apply_template_guidance(
        self,
        base_schema: TemplateSchema,
        guidance_schema: TemplateSchema | None,
    ) -> TemplateSchema:
        if guidance_schema is None:
            guidance_schema = TemplateSchema(workbook_filename=base_schema.workbook_filename, sheets=[])

        base_sheet_map = {sheet.name: sheet for sheet in base_schema.sheets}
        guidance_sheet_map = {sheet.name: sheet for sheet in guidance_schema.sheets}

        unknown_sheets = sorted(set(guidance_sheet_map) - set(base_sheet_map))
        if unknown_sheets:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Template guidance contains unknown sheet(s): {', '.join(unknown_sheets)}.",
            )

        merged_sheets: list[SheetSchema] = []
        for base_sheet in base_schema.sheets:
            guidance_sheet = guidance_sheet_map.get(base_sheet.name)
            guidance_column_map = {column.name: column for column in guidance_sheet.columns} if guidance_sheet else {}
            base_column_names = {column.name for column in base_sheet.columns}
            unknown_columns = sorted(set(guidance_column_map) - base_column_names)
            if unknown_columns:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Template guidance for sheet '{base_sheet.name}' contains unknown column(s): {', '.join(unknown_columns)}.",
                )

            merged_columns = [
                ColumnSchema(
                    name=column.name,
                    column_index=column.column_index,
                    excel_column=column.excel_column,
                    description=(guidance_column_map.get(column.name).description if guidance_column_map.get(column.name) else "").strip(),
                )
                for column in base_sheet.columns
            ]
            merged_sheets.append(
                SheetSchema(
                    name=base_sheet.name,
                    header_row_index=base_sheet.header_row_index,
                    data_start_row_index=base_sheet.data_start_row_index,
                    columns=merged_columns,
                )
            )

        return TemplateSchema(
            workbook_filename=base_schema.workbook_filename,
            sheets=merged_sheets,
        )

    def build_prompt_schema(self, template_schema: TemplateSchema) -> dict[str, dict[str, dict[str, str]]]:
        return {
            sheet.name: {
                column.name: {
                    "type": "string",
                    "guidance": column.description.strip() or "No extra column guidance provided.",
                }
                for column in sheet.columns
            }
            for sheet in template_schema.sheets
        }

    def _find_header_row_index(self, worksheet) -> int | None:
        for row_index in range(1, worksheet.max_row + 1):
            if any(cell.value is not None and str(cell.value).strip() for cell in worksheet[row_index]):
                return row_index
        return None

    def _find_first_empty_row(self, worksheet, sheet_schema: SheetSchema) -> int:
        row_index = sheet_schema.data_start_row_index
        upper_bound = max(worksheet.max_row, sheet_schema.data_start_row_index)
        while row_index <= upper_bound:
            has_value = any(
                worksheet.cell(row=row_index, column=column.column_index).value not in (None, "")
                for column in sheet_schema.columns
            )
            if not has_value:
                return row_index
            row_index += 1
        return row_index

    def _copy_row_style(self, worksheet, source_row: int, target_row: int, columns: list[ColumnSchema]) -> None:
        if source_row > worksheet.max_row:
            return

        worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
        for column in columns:
            source = worksheet.cell(row=source_row, column=column.column_index)
            target = worksheet.cell(row=target_row, column=column.column_index)
            if source.has_style:
                target._style = copy(source._style)
            target.number_format = source.number_format
            target.alignment = copy(source.alignment)
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.protection = copy(source.protection)

    def _resolve_cell_value(
        self,
        column_name: str,
        result: PaperExtractionResult,
        sheet_output: dict[str, str],
    ) -> str:
        if column_name in sheet_output and sheet_output[column_name]:
            return sheet_output[column_name]

        lowered = column_name.strip().lower()
        metadata_map = {
            "source_file": result.paper_filename,
            "paper_filename": result.paper_filename,
            "pdf_filename": result.paper_filename,
            "paper_id": result.paper_id,
            "sequence_number": str(result.sequence_number),
        }
        return metadata_map.get(lowered, "")


_template_service = TemplateService(get_repository())


def get_template_service() -> TemplateService:
    return _template_service
